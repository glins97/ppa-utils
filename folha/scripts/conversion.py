import openpyxl 
import datetime
import re
from copy import copy, deepcopy
import locale
import numpy
import subprocess

def load_data(fn):
    src = ''
    with open(fn, 'r') as f:    
        src = f.read()

    data = []
    for line in src.split('\n'):
        data.append(line.split(','))
    
    return data

def get_user_data(fn):
    data = load_data(fn)
    user_data = {}
    for id, date, user, enter_time, exit_time in data[:-1]:
        date_ = datetime.datetime.strptime(date, '%d/%m/%Y') 
        start_date = date_
        start_date = start_date.replace(hour=int(enter_time.split(':')[0]), minute=int(enter_time.split(':')[1]))
        
        end_date = date_
        end_date = end_date.replace(hour=int(exit_time.split(':')[0]), minute=int(exit_time.split(':')[1]))
        delta = ':'.join(str(end_date - start_date).split(':')[:-1])
        worked_hours = (end_date - start_date).seconds / 3600
        rest = worked_hours - int(worked_hours)
        if rest < 0.25:
            worked_hours = int(worked_hours)
        elif rest >= 0.25 and rest < 0.5:
            worked_hours = int(worked_hours) + 0.5
        elif rest >= 0.5 and rest < 0.75:
            worked_hours = int(worked_hours) + 0.75
        else:
            worked_hours = int(worked_hours) + 1
        delta_worked_hours = ':'.join(str(datetime.timedelta(hours=worked_hours)).split(':')[:-1])

        val = worked_hours * 26  
        if exit_time == '00:00':
            delta = '00:00'
            delta_worked_hours = '00:00'
            val = 0
        if user not in user_data:
            user_data[user] = [[date, enter_time, exit_time, delta, delta_worked_hours, val]]
        else:
            user_data[user].append([date, enter_time, exit_time, delta, delta_worked_hours, val])
    return user_data

def get_row(cell):
    return int(re.search(r'(\d+?)$', cell).group(1))
    
def duplicate(ws, origin, destination):
    ws[destination].font = copy(ws[origin].font)
    ws[destination].border = copy(ws[origin].border)
    ws[destination].fill = copy(ws[origin].fill)
    ws[destination].number_format = copy(ws[origin].number_format)
    ws[destination].protection = copy(ws[origin].protection)
    ws[destination].alignment = copy(ws[origin].alignment)
    if type(ws[destination]).__name__ != 'MergedCell':
        ws[destination].value = copy(ws[origin].value)

    if (ws.row_dimensions[get_row(origin)].height == 7.5):
        ws.row_dimensions[get_row(destination)].height = 7.5

def csv_to_xlsx(fn):
    user_data = get_user_data(fn)
    wb = openpyxl.load_workbook(filename='inputs/.TEMPLATE_FOLHA.xlsx')
    ws = wb.active
    blocked_users = ['Rafael Riemma', 'Manoel Neto', 'Kazuo', 'Breno Cunha', 'Raquel Oliveira']   
    count = 0
    ppa_total = 0
    for user in sorted(user_data.keys()):
        if user in blocked_users:
            continue
        user_total = 0
        for i, data in enumerate(user_data[user]):
            count += 1
            user_val = int(data[5])
            user_total += user_val
            row = str(4 + count)
            duplicate(ws, 'B4', 'B' + row)
            duplicate(ws, 'C4', 'C' + row)
            duplicate(ws, 'D4', 'D' + row)
            duplicate(ws, 'E4', 'E' + row)
            duplicate(ws, 'F4', 'F' + row)
            duplicate(ws, 'G4', 'G' + row)
            duplicate(ws, 'H4', 'H' + row)
            if i == 0:
                ws['B' + row] = '    ' + user
            ws['C' + row] = data[0]
            ws['D' + row] = data[1]
            ws['E' + row] = data[2]
            ws['F' + row] = data[3]
            ws['G' + row] = data[4]
            ws['H' + row] = 'R$ {:0,.0f}'.format(user_val).replace(',', '~').replace('.', ',').replace('~', '.')
            ws['B' + row].border = copy(ws['A1'].border)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['D' + row].border = copy(ws['A1'].border)
            ws['E' + row].border = copy(ws['A1'].border)
            ws['F' + row].border = copy(ws['A1'].border)
            ws['G' + row].border = copy(ws['A1'].border)
            ws['H' + row].border = copy(ws['A1'].border)
        count += 1
        row = str(4 + count)
        duplicate(ws, 'B4', 'B' + row)
        duplicate(ws, 'C4', 'C' + row)
        duplicate(ws, 'D4', 'D' + row)
        duplicate(ws, 'E4', 'E' + row)
        duplicate(ws, 'F4', 'F' + row)
        duplicate(ws, 'G4', 'G' + row)
        duplicate(ws, 'H4', 'H' + row)
        ws['H' + row] = 'R$ {:0,.0f}'.format(user_total).replace(',', '~').replace('.', ',').replace('~', '.')
        ws['B' + row].border = copy(ws['B4'].border)
        ws['C' + row].border = copy(ws['C4'].border)
        ws['D' + row].border = copy(ws['D4'].border)
        ws['E' + row].border = copy(ws['E4'].border)
        ws['F' + row].border = copy(ws['F4'].border)
        ws['G' + row].border = copy(ws['G4'].border)
        ws['H' + row].border = copy(ws['H4'].border)
        ppa_total += user_total
    count += 1
    row = str(4 + count)
    duplicate(ws, 'B4', 'B' + row)
    duplicate(ws, 'C4', 'C' + row)
    duplicate(ws, 'D4', 'D' + row)
    duplicate(ws, 'E4', 'E' + row)
    duplicate(ws, 'F4', 'F' + row)
    duplicate(ws, 'G4', 'G' + row)
    ws['B' + row] = '    TOTAL PPA'
    ws['H' + row] = 'R$ {:0,.0f}'.format(ppa_total).replace(',', '~').replace('.', ',').replace('~', '.')
    ws['B' + row].border = copy(ws['B4'].border)
    ws['C' + row].border = copy(ws['C4'].border)
    ws['D' + row].border = copy(ws['D4'].border)
    ws['E' + row].border = copy(ws['E4'].border)
    ws['F' + row].border = copy(ws['F4'].border)
    ws['G' + row].border = copy(ws['G4'].border)
    ws['H' + row].border = copy(ws['H4'].border)
    ws.delete_rows(4, 1)

    out = fn.replace('.csv', '.xlsx')
    wb.save(out)
    return out

def xlsx_to_pdf(fn):
    subprocess.call(['libreoffice', '--headless', '--convert-to',  'pdf', fn, '--outdir', 'outputs'])
    return fn.replace('inputs', 'outputs').replace('xlsx', 'pdf')

if __name__ == "__main__":
    main()